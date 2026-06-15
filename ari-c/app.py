import io
import os
import zipfile
import tempfile
from datetime import datetime

import cv2
import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font as XLFont, Alignment, PatternFill, Border, Side
from PIL import Image

from inference import C2C7Inference
from cobb_draw import draw_cobb_angle

MAX_FILE_SIZE_MB = 10
MAX_FILE_SIZE_BYTES = MAX_FILE_SIZE_MB * 1024 * 1024

st.set_page_config(
    page_title="ARI-C · C-spine Lateral Radiograph Analyzer",
    page_icon="🦴",
    layout="wide",
)

# ─── Apple Product Page dark theme ──────────────────────────────────────────
st.markdown("""
<style>
body, .stApp {
    background: #1d1d1f !important;
    font-family: -apple-system, BlinkMacSystemFont, 'SF Pro Display', 'Helvetica Neue', sans-serif;
}
.main .block-container {
    padding: 68px 2.5rem 5rem;
    max-width: 980px;
}
#MainMenu { visibility: hidden; }
footer { visibility: hidden; }
header { visibility: hidden; }
[data-testid="stDecoration"] { display: none !important; }
[data-testid="stToolbar"] { display: none !important; }

/* ─── Fixed navbar ── */
.ari-nav {
    position: fixed;
    top: 0; left: 0; right: 0;
    z-index: 9999;
    height: 44px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 0 28px;
    background: rgba(29,29,31,.78);
    backdrop-filter: saturate(180%) blur(20px);
    -webkit-backdrop-filter: saturate(180%) blur(20px);
    border-bottom: 1px solid rgba(255,255,255,.06);
}
.ari-nav-back {
    display: inline-flex;
    align-items: center;
    gap: 5px;
    font-size: 14px;
    font-weight: 400;
    color: #2ec4e8;
    text-decoration: none;
    letter-spacing: -.01em;
    transition: opacity .15s;
}
.ari-nav-back:hover { opacity: .65; }
.ari-nav-back svg {
    flex-shrink: 0;
    margin-top: 1px;
}
.ari-nav-brand {
    font-size: 13px;
    font-weight: 500;
    color: rgba(255,255,255,.4);
    letter-spacing: .02em;
    text-transform: uppercase;
}

/* ─── Page header ── */
.ari-header {
    padding: 52px 0 36px;
    border-bottom: 1px solid rgba(255,255,255,.07);
    margin-bottom: 40px;
}
.ari-eyebrow {
    font-size: 11px;
    font-weight: 600;
    letter-spacing: .14em;
    text-transform: uppercase;
    color: #2ec4e8;
    margin: 0 0 16px;
}
.ari-title {
    font-size: 52px;
    font-weight: 700;
    letter-spacing: -.03em;
    color: #f5f5f7;
    margin: 0 0 14px;
    line-height: 1.0;
}
.ari-sub {
    font-size: 19px;
    color: #86868b;
    font-weight: 400;
    letter-spacing: -.01em;
    margin: 0 0 24px;
    line-height: 1.5;
}
.ari-specs {
    display: flex;
    align-items: center;
    gap: 8px;
    flex-wrap: wrap;
}
.ari-chip {
    font-size: 12px;
    font-weight: 500;
    color: #98989d;
    background: rgba(255,255,255,.05);
    border: 1px solid rgba(255,255,255,.08);
    border-radius: 99px;
    padding: 4px 13px;
    letter-spacing: .01em;
    white-space: nowrap;
}

/* ─── Section labels ── */
.ari-section {
    font-size: 22px;
    font-weight: 600;
    color: #f5f5f7;
    letter-spacing: -.015em;
    margin: 44px 0 16px;
    padding-bottom: 12px;
    border-bottom: 1px solid rgba(255,255,255,.06);
}

/* ─── File uploader ── */
[data-testid="stFileUploader"] {
    background: rgba(255,255,255,.03) !important;
    border: 1.5px dashed rgba(255,255,255,.13) !important;
    border-radius: 16px !important;
    padding: 8px !important;
}
[data-testid="stFileUploader"]:focus-within {
    border-color: rgba(46,196,232,.45) !important;
}
[data-testid="stFileUploaderDropzone"],
[data-testid="stFileUploader"] section {
    background: transparent !important;
    border: none !important;
}
[data-testid="stFileUploader"] label { color: #f5f5f7 !important; font-weight: 500 !important; }
[data-testid="stFileUploader"] small { color: #6e6e73 !important; }
[data-testid="stFileUploader"] button {
    background: rgba(46,196,232,.1) !important;
    color: #2ec4e8 !important;
    border: 1px solid rgba(46,196,232,.22) !important;
    border-radius: 10px !important;
    font-weight: 500 !important;
    font-size: 14px !important;
}
[data-testid="stFileUploader"] button:hover {
    background: rgba(46,196,232,.18) !important;
}

/* ─── Analyze button ── */
.stButton > button[kind="primary"] {
    background: #2ec4e8 !important;
    color: #000 !important;
    border: none !important;
    border-radius: 12px !important;
    font-weight: 600 !important;
    font-size: 15px !important;
    padding: 10px 32px !important;
    box-shadow: 0 0 24px rgba(46,196,232,.25) !important;
}
.stButton > button[kind="primary"]:hover { opacity: .85 !important; }

/* ─── Download buttons ── */
.stDownloadButton > button {
    background: rgba(255,255,255,.05) !important;
    color: #f5f5f7 !important;
    border: 1px solid rgba(255,255,255,.1) !important;
    border-radius: 10px !important;
    font-weight: 500 !important;
    padding: 9px 20px !important;
}
.stDownloadButton > button:hover { background: rgba(255,255,255,.09) !important; }

/* ─── Progress bar ── */
[data-testid="stProgressBar"] > div {
    background: rgba(255,255,255,.08) !important;
    border-radius: 99px !important;
    height: 4px !important;
}
[data-testid="stProgressBar"] > div > div {
    background: linear-gradient(90deg, #2ec4e8, #0071e3) !important;
    border-radius: 99px !important;
}

/* ─── Dataframe ── */
[data-testid="stDataFrame"] {
    border-radius: 12px !important;
    overflow: hidden !important;
    border: 1px solid rgba(255,255,255,.06) !important;
}

/* ─── Images ── */
[data-testid="stImage"] img {
    border-radius: 10px !important;
    border: 1px solid rgba(255,255,255,.07) !important;
}
[data-testid="stImage"] p {
    color: #6e6e73 !important;
    font-size: 12px !important;
    text-align: center !important;
}

/* ─── Alerts ── */
[data-testid="stError"] {
    background: rgba(255,69,58,.07) !important;
    border: 1px solid rgba(255,69,58,.2) !important;
    border-radius: 12px !important;
}
[data-testid="stWarning"] {
    background: rgba(255,159,10,.07) !important;
    border: 1px solid rgba(255,159,10,.2) !important;
    border-radius: 12px !important;
}

/* ─── Typography ── */
h2, h3 { color: #f5f5f7 !important; font-weight: 600 !important; letter-spacing: -.015em !important; }
hr { border-color: rgba(255,255,255,.07) !important; margin: 32px 0 !important; }
.stCaption { color: #6e6e73 !important; }

/* ─── Status summary ── */
.ari-status-summary {
    margin-bottom: 20px;
    border-radius: 14px;
    overflow: hidden;
    border: 1px solid rgba(255,255,255,.07);
}
.ari-status-header {
    display: flex;
    align-items: center;
    gap: 10px;
    padding: 12px 20px;
    font-size: 14px;
    font-weight: 600;
    letter-spacing: -.01em;
}
.ari-status-header.all-ok {
    background: rgba(48,209,88,.08);
    border-bottom: 1px solid rgba(48,209,88,.15);
    color: #30d158;
}
.ari-status-header.has-err {
    background: rgba(255,159,10,.08);
    border-bottom: 1px solid rgba(255,159,10,.15);
    color: #ff9f0a;
}
.ari-status-rows { padding: 6px 0; }
.ari-status-row {
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 7px 20px;
    font-size: 13px;
    border-bottom: 1px solid rgba(255,255,255,.04);
}
.ari-status-row:last-child { border-bottom: none; }
.ari-status-fname {
    color: #98989d;
    flex: 1;
    min-width: 0;
    overflow: hidden;
    text-overflow: ellipsis;
    white-space: nowrap;
    margin-right: 16px;
}
.ari-badge {
    flex-shrink: 0;
    font-size: 11px;
    font-weight: 600;
    padding: 3px 10px;
    border-radius: 99px;
    letter-spacing: .03em;
}
.ari-badge.ok  { background: rgba(48,209,88,.12);  color: #30d158; border: 1px solid rgba(48,209,88,.2); }
.ari-badge.err { background: rgba(255,69,58,.12);  color: #ff453a; border: 1px solid rgba(255,69,58,.2); }

/* ─── Scrollbar ── */
::-webkit-scrollbar { width: 5px; height: 5px; }
::-webkit-scrollbar-track { background: #1d1d1f; }
::-webkit-scrollbar-thumb { background: rgba(255,255,255,.12); border-radius: 99px; }
::-webkit-scrollbar-thumb:hover { background: rgba(255,255,255,.22); }

/* ─── Version cards ── */
.ari-cards {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 16px;
    margin: 0 0 40px;
}
.ari-card {
    background: rgba(255,255,255,.03);
    border: 1px solid rgba(255,255,255,.08);
    border-radius: 18px;
    padding: 28px 28px 24px;
}
.ari-card-title {
    font-size: 18px;
    font-weight: 600;
    color: #f5f5f7;
    letter-spacing: -.015em;
    margin: 0 0 18px;
}
.ari-card ul {
    list-style: none;
    margin: 0;
    padding: 0;
    display: flex;
    flex-direction: column;
    gap: 10px;
}
.ari-card li {
    font-size: 14px;
    color: #98989d;
    display: flex;
    align-items: flex-start;
    gap: 8px;
    line-height: 1.45;
}
.ari-card li strong { color: #d1d1d6; }
.ari-card-video {
    margin-top: 22px;
    border-radius: 12px;
    overflow: hidden;
    position: relative;
    width: 100%;
    aspect-ratio: 16 / 9;
    background: #000;
}
.ari-card-video iframe {
    position: absolute;
    inset: 0;
    width: 100%;
    height: 100%;
    border: 0;
}
.ari-card-contact {
    margin-top: 16px;
    font-size: 13px;
    color: #6e6e73;
}
.ari-card-contact a { color: #2ec4e8; text-decoration: none; }
.ari-card-contact a:hover { text-decoration: underline; }
.ari-card-btn {
    display: flex;
    align-items: center;
    justify-content: center;
    gap: 7px;
    margin-top: 22px;
    padding: 11px 20px;
    background: #2ec4e8;
    color: #000 !important;
    border-radius: 11px;
    font-size: 15px;
    font-weight: 600;
    text-decoration: none !important;
    letter-spacing: -.01em;
    transition: opacity .15s;
    box-sizing: border-box;
    width: 100%;
}
.ari-card-btn:hover { opacity: .82; }
@media (max-width: 640px) {
    .ari-cards { grid-template-columns: 1fr; }
}

/* ─── Standalone list (outside .ari-card) ── */
.ari-card-list {
    list-style: none;
    margin: 0 0 16px;
    padding: 0;
    display: flex;
    flex-direction: column;
    gap: 10px;
}
.ari-card-list li {
    font-size: 14px;
    color: #98989d;
    display: flex;
    align-items: flex-start;
    gap: 8px;
    line-height: 1.45;
}
.ari-card-list li strong { color: #d1d1d6; }

/* ─── Web Version merged container ── */
[data-testid="stVerticalBlockBorderWrapper"] {
    background: rgba(255,255,255,.03) !important;
    border: 1px solid rgba(255,255,255,.08) !important;
    border-radius: 18px !important;
}
[data-testid="stVerticalBlockBorderWrapper"] > div {
    padding: 16px 16px 12px !important;
}
</style>
""", unsafe_allow_html=True)

# ─── Navbar + Header ─────────────────────────────────────────────────────────
st.markdown(f"""
<nav class="ari-nav">
  <a href="https://ariagent.co.kr" class="ari-nav-back">
    <svg width="8" height="13" viewBox="0 0 8 13" fill="none" xmlns="http://www.w3.org/2000/svg">
      <path d="M7 1L1 6.5L7 12" stroke="#2ec4e8" stroke-width="1.6" stroke-linecap="round" stroke-linejoin="round"/>
    </svg>
    ARI Agent
  </a>
  <span class="ari-nav-brand">ARI-C</span>
</nav>

<div class="ari-header">
  <p class="ari-eyebrow">ARI-C &nbsp;·&nbsp; C-Spine Module</p>
  <h1 class="ari-title">C-spine Lateral<br>Radiograph Analyzer</h1>
  <p class="ari-sub">C2–C7 Cobb Angle &nbsp;·&nbsp; C2 Sagittal Vertical Axis &nbsp;·&nbsp; C7 Slope</p>
  <div class="ari-specs">
  </div>
</div>
""", unsafe_allow_html=True)

_col_d, _col_w = st.columns(2, gap="medium")

with _col_d:
    st.markdown("""
<div class="ari-card">
  <p class="ari-card-title">💻 On-Device Desktop Version</p>
  <ul>
    <li>✅&nbsp; No internet connection required</li>
    <li>✅&nbsp; No installation required — runs as a standalone app</li>
    <li>✅&nbsp; Supports <strong>DICOM · JPG · PNG · BMP</strong> and <strong>clipboard paste</strong></li>
    <li>✅&nbsp; <strong>Unlimited</strong> file analysis</li>
    <li>⚡&nbsp; Batch throughput: ~<strong>10,000 images in 4 hours</strong></li>
  </ul>
  <div class="ari-card-video">
    <iframe src="https://www.youtube.com/embed/25_CxOHhRh8"
      title="ARI-C Demo Video"
      allow="accelerometer; autoplay; clipboard-write; encrypted-media; gyroscope; picture-in-picture; web-share"
      allowfullscreen></iframe>
  </div>
  <p class="ari-card-contact">Purchase inquiry: <a href="mailto:imspinesurgeon@gmail.com">imspinesurgeon@gmail.com</a></p>
</div>
""", unsafe_allow_html=True)

with _col_w:
    with st.container(border=True):
        st.markdown(f"""
<p class="ari-card-title">🌐 Web Version</p>
<ul class="ari-card-list">
  <li>🌐&nbsp; Accessible from any device via browser</li>
  <li>📁&nbsp; Supports <strong>DICOM · JPG · PNG · BMP</strong></li>
  <li>⚠️&nbsp; File size limit: <strong>{MAX_FILE_SIZE_MB} MB per upload</strong></li>
  <li>📊&nbsp; Auto-analysis · Excel &amp; ZIP download</li>
</ul>
""", unsafe_allow_html=True)
        uploaded_files = st.file_uploader(
            "Upload image files",
            type=["jpg", "jpeg", "png", "bmp", "dcm"],
            accept_multiple_files=True,
        )


@st.cache_resource
def load_engine():
    return C2C7Inference()


def pil_to_cv2(pil_img):
    img = np.array(pil_img)
    if img.ndim == 2:
        return cv2.cvtColor(img, cv2.COLOR_GRAY2BGR)
    if img.shape[2] == 4:
        return cv2.cvtColor(img, cv2.COLOR_RGBA2BGR)
    return cv2.cvtColor(img, cv2.COLOR_RGB2BGR)


def cv2_to_pil(cv_img):
    if cv_img.ndim == 2:
        return Image.fromarray(cv_img)
    rgb = cv2.cvtColor(cv_img, cv2.COLOR_BGR2RGB)
    return Image.fromarray(rgb)


if uploaded_files:
    total_size = sum(f.size for f in uploaded_files)
    oversized_files = [f.name for f in uploaded_files if f.size > MAX_FILE_SIZE_BYTES]

    if total_size > MAX_FILE_SIZE_BYTES:
        st.error(f"**File size limit exceeded!** Total: **{total_size / (1024*1024):.1f}MB** (Limit: {MAX_FILE_SIZE_MB}MB)")
        if oversized_files:
            st.warning(f"Oversized files: {', '.join(oversized_files)}")
        st.markdown("""
        ---
        **Desktop version** has no file size limits, supports batch processing of thousands of files, and works offline.

        Contact: [imspinesurgeon@gmail.com](mailto:imspinesurgeon@gmail.com)
        """)
        st.stop()

    # Auto-analyze when file selection changes
    _file_key = tuple(sorted((f.name, f.size) for f in uploaded_files))
    if st.session_state.get("_last_file_key") != _file_key:
        st.session_state["_last_file_key"] = _file_key
        engine = load_engine()
        all_results = []
        result_images = {}

        progress_bar = st.progress(0, text="Processing...")
        total = len(uploaded_files)

        for idx, uploaded_file in enumerate(uploaded_files):
            filename = uploaded_file.name
            progress_bar.progress(idx / total, text=f"Processing: {filename} ({idx+1}/{total})")

            try:
                suffix = os.path.splitext(filename)[1]
                with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                    tmp.write(uploaded_file.getbuffer())
                    tmp_path = tmp.name

                results_list, original_color, regions = engine.run_multi(tmp_path)
                os.unlink(tmp_path)
                num_regions = len(regions)

                if num_regions == 1:
                    result_tuple = results_list[0]
                    if len(result_tuple) == 6:
                        measurements, keypoints, sub_img, dicom_info, region_idx, error_msg = result_tuple
                        all_results.append({
                            "Filename": filename,
                            "C2-C7 Angle (°)": "", "C2 Slope (°)": "", "C7 Slope (°)": "",
                            "C2-C7 SVA (mm)": "", "C2-C7 SVA (px)": "",
                            "Sex": "", "Age": "", "Study Date": "", "Study ID": "", "Study Desc": "",
                            "Status": f"Error: {error_msg[:60]}",
                        })
                    else:
                        measurements, keypoints, sub_img, dicom_info, region_idx = result_tuple
                        result_img = draw_cobb_angle(original_color.copy(), keypoints, measurements)
                        sva_mm = measurements.get("sva_mm") if measurements.get("sva_mm") is not None else ""
                        dicom_info = dicom_info or {}
                        all_results.append({
                            "Filename": filename,
                            "C2-C7 Angle (°)": measurements["cobb_angle"],
                            "C2 Slope (°)": measurements["c2_slope"],
                            "C7 Slope (°)": measurements["c7_slope"],
                            "C2-C7 SVA (mm)": sva_mm,
                            "C2-C7 SVA (px)": measurements["sva_px"],
                            "Sex": dicom_info.get("sex", ""),
                            "Age": dicom_info.get("age", ""),
                            "Study Date": dicom_info.get("study_date", ""),
                            "Study ID": dicom_info.get("study_id", ""),
                            "Study Desc": dicom_info.get("study_description", ""),
                            "Status": "Completed",
                        })
                        result_images[filename] = {"original": original_color, "result": result_img}
                else:
                    combined_result = original_color.copy()
                    for result_tuple in results_list:
                        if len(result_tuple) == 6:
                            measurements, keypoints, sub_img, dicom_info, region_idx, error_msg = result_tuple
                            all_results.append({
                                "Filename": filename + f" (Region {region_idx + 1})",
                                "C2-C7 Angle (°)": "", "C2 Slope (°)": "", "C7 Slope (°)": "",
                                "C2-C7 SVA (mm)": "", "C2-C7 SVA (px)": "",
                                "Sex": "", "Age": "", "Study Date": "", "Study ID": "", "Study Desc": "",
                                "Status": f"Error: {error_msg[:60]}",
                            })
                            continue
                        measurements, keypoints, sub_img, dicom_info, region_idx = result_tuple
                        sub_result = draw_cobb_angle(sub_img.copy(), keypoints, measurements)
                        x_start, x_end = regions[region_idx]
                        combined_result[:, x_start:x_end] = sub_result
                        sva_mm = measurements.get("sva_mm") if measurements.get("sva_mm") is not None else ""
                        dicom_info = dicom_info or {}
                        all_results.append({
                            "Filename": filename + f" (Region {region_idx + 1})",
                            "C2-C7 Angle (°)": measurements["cobb_angle"],
                            "C2 Slope (°)": measurements["c2_slope"],
                            "C7 Slope (°)": measurements["c7_slope"],
                            "C2-C7 SVA (mm)": sva_mm,
                            "C2-C7 SVA (px)": measurements["sva_px"],
                            "Sex": dicom_info.get("sex", ""),
                            "Age": dicom_info.get("age", ""),
                            "Study Date": dicom_info.get("study_date", ""),
                            "Study ID": dicom_info.get("study_id", ""),
                            "Study Desc": dicom_info.get("study_description", ""),
                            "Status": "Completed",
                        })
                    result_images[filename] = {"original": original_color, "result": combined_result}

            except Exception as e:
                all_results.append({
                    "Filename": filename,
                    "C2-C7 Angle (°)": "", "C2 Slope (°)": "", "C7 Slope (°)": "",
                    "C2-C7 SVA (mm)": "", "C2-C7 SVA (px)": "",
                    "Sex": "", "Age": "", "Study Date": "", "Study ID": "", "Study Desc": "",
                    "Status": f"Error: {str(e)[:80]}",
                })
                if 'tmp_path' in locals() and os.path.exists(tmp_path):
                    os.unlink(tmp_path)

        progress_bar.progress(1.0, text="Processing complete!")
        st.session_state["all_results"] = all_results
        st.session_state["result_images"] = result_images

    if "all_results" in st.session_state:
        all_results = st.session_state["all_results"]
        result_images = st.session_state["result_images"]

        # ── Status summary ─────────────────────────────────────────────────────
        n_total = len(all_results)
        n_ok  = sum(1 for r in all_results if r["Status"] == "Completed")
        n_err = n_total - n_ok

        header_cls = "all-ok" if n_err == 0 else "has-err"
        header_icon = "✓" if n_err == 0 else "⚠"
        if n_err == 0:
            header_text = f"{n_ok} / {n_total} &nbsp; Completed"
        else:
            header_text = (
                f"{n_ok} / {n_total} &nbsp; Completed"
                + (f" &nbsp;·&nbsp; {n_err} Error{'s' if n_err > 1 else ''}" if n_err else "")
            )

        rows_html = ""
        for r in all_results:
            fname = r["Filename"]
            status = r["Status"]
            if status == "Completed":
                badge = '<span class="ari-badge ok">Completed</span>'
            else:
                short = status[:60] + "…" if len(status) > 60 else status
                badge = f'<span class="ari-badge err">{short}</span>'
            rows_html += (
                f'<div class="ari-status-row">'
                f'<span class="ari-status-fname">{fname}</span>'
                f'{badge}'
                f'</div>'
            )

        st.markdown(
            f'<div class="ari-status-summary">'
            f'<div class="ari-status-header {header_cls}">{header_icon}&nbsp; {header_text}</div>'
            f'<div class="ari-status-rows">{rows_html}</div>'
            f'</div>',
            unsafe_allow_html=True,
        )

        # ── Results table ──────────────────────────────────────────────────────
        st.markdown('<p class="ari-section">Measurement Results</p>', unsafe_allow_html=True)
        df = pd.DataFrame(all_results)
        st.dataframe(df, use_container_width=True, hide_index=True)

        # ── Result images ──────────────────────────────────────────────────────
        if result_images:
            st.markdown('<p class="ari-section">Result Images</p>', unsafe_allow_html=True)
            for filename, data in result_images.items():
                st.markdown(f"**{filename}**")
                col1, col2 = st.columns(2)
                with col1:
                    st.image(cv2_to_pil(data["original"]), caption="Original", use_container_width=True)
                with col2:
                    st.image(cv2_to_pil(data["result"]), caption="Result", use_container_width=True)

        # ── Downloads ─────────────────────────────────────────────────────────
        st.markdown('<p class="ari-section">Downloads</p>', unsafe_allow_html=True)
        col_dl1, col_dl2 = st.columns(2)

        with col_dl1:
            wb = Workbook()
            ws = wb.active
            ws.title = "C2-C7 Measurements"
            header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
            header_font = XLFont(bold=True, size=11, color="2EC4E8")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            center_align = Alignment(horizontal='center', vertical='center')
            headers = ["Filename", "C2-C7 Angle (°)", "C2 Slope (°)",
                       "C7 Slope (°)", "C2-C7 SVA (mm)", "C2-C7 SVA (px)",
                       "Sex", "Age", "Study Date", "Study ID", "Study Desc", "Status"]
            ws.append(headers)
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            for row in all_results:
                ws.append([
                    row["Filename"], row["C2-C7 Angle (°)"], row["C2 Slope (°)"],
                    row["C7 Slope (°)"], row["C2-C7 SVA (mm)"], row["C2-C7 SVA (px)"],
                    row.get("Sex", ""), row.get("Age", ""), row.get("Study Date", ""),
                    row.get("Study ID", ""), row.get("Study Desc", ""), row["Status"],
                ])
            for row_idx in range(2, ws.max_row + 1):
                for col in range(1, 13):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.border = thin_border
                    if col > 1:
                        cell.alignment = center_align
            ws.column_dimensions['A'].width = 30
            for col_letter in ['B', 'C', 'D', 'E', 'F']:
                ws.column_dimensions[col_letter].width = 16
            ws.column_dimensions['G'].width = 8
            ws.column_dimensions['H'].width = 8
            ws.column_dimensions['I'].width = 12
            ws.column_dimensions['J'].width = 12
            ws.column_dimensions['K'].width = 20
            ws.column_dimensions['L'].width = 20
            excel_buf = io.BytesIO()
            wb.save(excel_buf)
            excel_buf.seek(0)
            timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
            st.download_button(
                label="Download Excel",
                data=excel_buf,
                file_name=f"C-lat_angle_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        with col_dl2:
            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                for filename, data in result_images.items():
                    basename = os.path.splitext(filename)[0]
                    _, png_bytes = cv2.imencode(".png", data["result"])
                    zf.writestr(f"result_{basename}.png", png_bytes.tobytes())
            zip_buf.seek(0)
            st.download_button(
                label="Download Result Images (ZIP)",
                data=zip_buf,
                file_name="c2c7_results.zip",
                mime="application/zip",
            )
